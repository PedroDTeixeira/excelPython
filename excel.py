import openpyxl
wb = openpyxl.load_workbook('SalesFigures.xlsx')
sheet = wb['Sheet1']
exampleOne = sheet['A1']
exampleTwo = sheet['A1'].value
c = sheet['B1']
# print(c.value, c.row, c.column, c.coordinate)
# print(sheet.cell(row=1, column=2).value)
# for i in range(1, 8, 2):
#     print(i,sheet.cell(row=i, column=2).value)
# print(sheet.max_column)
# print(sheet.max_row)

# tuple(sheet['A1':'C3'])

# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')

# list(sheet.columns)[1]

# for cellObj in list(sheet.columns)[1]:
#     print(cellObj.value)
